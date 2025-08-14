import os
import re
import shutil
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

class KATOFileSorter:
    """
    Класс для сортировки, объединения и экспорта Excel-файлов с кодами КАТО по регионам и приложениям.

    Основные функции:
    - Поиск Excel-файлов во входной директории.
    - Удаление ранее созданных подпапок 'като_файлы'.
    - Извлечение данных по КАТО, группировка по 6-значному коду и приложению.
    - Удаление дубликатов записей.
    - Сохранение итоговых файлов в разрезе КАТО и региона в виде отдельных Excel-файлов.

    Атрибуты:
    -----------
    input_dir : str
        Путь к директории, где находятся исходные Excel-файлы.
    output_dir : str
        Путь для сохранения обработанных файлов.
    kato_data : defaultdict
        Словарь для хранения сгруппированных данных вида kato_data[kato_code][app_title] = list of (title, DataFrame).
    kato_skip_list : set
        Множество КАТО-кодов, которые нужно исключить из многократной обработки (например, 25022025).
    already_saved_kato : set
        Множество КАТО-кодов, по которым уже были сохранены итоговые файлы.
    added_entries : set
        Хеш-множество уникальных записей для защиты от дублирования (tuple: kato, app, filename).

    Методы:
    --------
    delete_kato_subfolders()
        Удаляет все подпапки с названием, начинающимся на 'като_файлы' во входной директории.
    
    get_excel_files() -> list
        Рекурсивно ищет все `.xlsx` файлы в input_dir и возвращает их пути.

    make_unique_columns(columns: list) -> list
        Делает названия колонок уникальными, добавляя суффиксы (_1, _2...) при повторениях.

    combine_column(col_tuple: tuple) -> str
        Объединяет кортежи названий колонок из многоуровневого заголовка в одну строку.

    extract_app_number(name: str) -> int
        Извлекает номер приложения из названия строки или файла. Используется для сортировки.

    get_region_folder(kato_code: str) -> str
        Определяет название области на основе первых двух цифр КАТО-кода.

    process_files()
        Основной метод для обработки всех Excel-файлов:
        - ищет приложения и заголовки,
        - извлекает таблицы и КАТО,
        - группирует по 6-значному корню КАТО и приложению,
        - предотвращает дублирование данных.

    save_kato_files()
        Сохраняет сгруппированные данные по КАТО в Excel-файлы по регионам.
        Каждое приложение добавляется как отдельный лист с форматированием заголовков и автошириной колонок.
    """
    MAX_WIDTH = 30
    REGION_CODES = {
        "11": "Акмолинская", "15": "Актюбинская", "19": "Алматинская",
        "23": "Атырауская", "63": "Восточно-Казахстанская", "71": "Астана",
        "75": "Алматы", "79": "Шымкент", "31": "Жамбылская",
        "27": "Западно-Казахстанская", "35": "Карагандинская", "39": "Костанайская",
        "43": "Кызылордиснкая", "47": "Мангистауская", "10": "область Абай",
        "33": "область Жетісу", "62": "область Ұлытау", "55": "Павлодарская",
        "59": "Северо-Казахстанская", "61": "Туркестанская"
    }

    def __init__(self, input_dir='.', output_dir='итоговые_файлы'):
        self.input_dir = input_dir
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        self.kato_data = defaultdict(lambda: defaultdict(list))
        self.kato_skip_list = {'25022025'}
        self.already_saved_kato = set()
        self.added_entries = set()

    def delete_kato_subfolders(self):
        deleted = 0
        for root, dirs, _ in os.walk(self.input_dir):
            for d in dirs:
                if d.lower().startswith("като_файлы"):
                    full_path = os.path.join(root, d)
                    try:
                        shutil.rmtree(full_path)
                        print(f"🗑 Удалена папка: {full_path}")
                        deleted += 1
                    except Exception as e:
                        print(f"⚠️ Не удалось удалить {full_path}: {e}")
        print(f"\n✅ Удалено {deleted} старых папок 'като_файлы'")

    def make_unique_columns(self, columns):
        seen = {}
        new_columns = []
        for col in columns:
            col = col.strip()
            if col in seen:
                seen[col] += 1
                new_columns.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                new_columns.append(col)
        return new_columns

    def combine_column(self, col_tuple):
        parts = [str(x).strip() for x in col_tuple if str(x).strip().lower() != 'nan' and 'Unnamed' not in str(x)]
        return ' / '.join(parts)

    def extract_app_number(self, name):
        match = re.search(r'(\d{1,2})', name)
        return int(match.group(1)) if match else 9999

    def get_region_folder(self, kato_code):
        region_prefix = kato_code[:2]
        return self.REGION_CODES.get(region_prefix, 'Прочее')

    def get_excel_files(self):
        excel_files = []
        for root, _, files in os.walk(self.input_dir):
            for file in files:
                if file.lower().endswith('.xlsx'):
                    excel_files.append(os.path.join(root, file))
        return excel_files

    def process_files(self):
        self.delete_kato_subfolders()
        all_excel_files = self.get_excel_files()
        print(f"\n🔍 Найдено Excel-файлов: {len(all_excel_files)}")

        for file_path in all_excel_files:
            file = os.path.basename(file_path)
            print(f"\n📂 Обрабатывается файл: {file}")

            try:
                preview = pd.read_excel(file_path, header=None, nrows=15)
                app_idx = None
                header_idx = None

                for i, row in preview.iterrows():
                    line = ' '.join([str(cell) for cell in row if pd.notna(cell)]).lower()
                    if app_idx is None and re.search(r'(\d{1,2})[- ]*(қосымша|kosymsha|приложение)', line):
                        app_idx = i
                    if header_idx is None and any("кaто" in str(cell).lower() or "kato" in str(cell).lower() for cell in row):
                        header_idx = i
                    if app_idx is not None and header_idx is not None:
                        break

                app_number = self.extract_app_number(file)
                app_title = f"Приложение-{app_number}"
                title_row = app_title

                if header_idx is not None:
                    if app_idx is not None and app_idx + 1 < header_idx:
                        title_row = ' '.join([str(cell) for cell in preview.iloc[app_idx + 1] if pd.notna(cell)])
                    df = pd.read_excel(file_path, header=header_idx, dtype=str)
                    df.columns = self.make_unique_columns([str(col) for col in df.columns])
                else:
                    print("⚠️ Переход в режим многоуровневого заголовка")
                    with pd.ExcelFile(file_path) as xls:
                        top_row = pd.read_excel(xls, header=None, nrows=1).iloc[0, 0]
                    title_row = str(top_row)
                    df = pd.read_excel(file_path, header=[1, 2, 3], engine='openpyxl', dtype=str)
                    df.columns = [self.combine_column(col) for col in df.columns]
                    df.columns = self.make_unique_columns(df.columns)

                df = df.dropna(how='all')
                df = df[~df.apply(lambda row: all(str(cell).strip().lower() in ['filt', 'nan', 'none', ''] for cell in row), axis=1)]

                kato_col_candidates = [col for col in df.columns if "kato" in col.lower()]
                if not kato_col_candidates:
                    kato1_col = [col for col in df.columns if "КАТО" in col]
                    if not kato1_col:
                        print(f"⚠️ Нет колонки КАТО в файле: {file}")
                        continue
                    kato_col_candidates.append(kato1_col[0])

                kato_col = kato_col_candidates[0]
                df[kato_col] = df[kato_col].astype(str).str.strip()
                df = df[df[kato_col].notna() & (df[kato_col].str.strip() != '')]
                df['kato_root'] = df[kato_col].str[:6]

                for kato_root, group_df in df.groupby('kato_root'):
                    if not kato_root.isdigit():
                        continue

                    entry_id = (kato_root, app_title, os.path.basename(file_path))
                    if entry_id in self.added_entries:
                        continue
                    self.added_entries.add(entry_id)
                    
                    self.kato_data[kato_root][app_title].append((title_row, group_df.drop(columns=['kato_root'])))

                print(f"✅ Успешно обработан: {file}")

            except Exception as e:
                print(f"❌ Ошибка при обработке файла {file}: {e}")

    def save_kato_files(self):
        for kato_code, app_dict in self.kato_data.items():
            if kato_code in self.kato_skip_list and kato_code in self.already_saved_kato:
                print(f"⏭ Пропущен повторяющийся КАТО: {kato_code}")
                continue

            region_name = self.get_region_folder(kato_code)
            region_folder = os.path.join(self.output_dir, region_name)
            os.makedirs(region_folder, exist_ok=True)

            wb = Workbook()
            wb.remove(wb.active)
            sorted_apps = sorted(app_dict.items(), key=lambda x: self.extract_app_number(x[0]))

            for app_title, entries in sorted_apps:
                sheet_name = app_title[:31]
                ws = wb.create_sheet(title=sheet_name)

                for title_row, df in entries:
                    ws.append([title_row])
                    ws.append([app_title])
                    ws["A1"].font = Font(bold=True, size=14)
                    ws["A1"].alignment = Alignment(horizontal='left')
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

                    ws["A2"].font = Font(bold=True)
                    ws["A2"].alignment = Alignment(horizontal='left')
                    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df.columns))

                    for i, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
                        ws.append(row)
                        if i == 3:
                            for cell in ws[i]:
                                cell.font = Font(bold=True)
                                cell.alignment = Alignment(horizontal='center')

                    for i, col in enumerate(ws.columns, start=1):
                        max_length = 0
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, self.MAX_WIDTH)
                        col_letter = get_column_letter(i)
                        ws.column_dimensions[col_letter].width = adjusted_width

            filename = f"{kato_code}.xlsx"
            output_path = os.path.join(region_folder, filename)
            wb.save(output_path)
            print(f"📄 Сохранён: {output_path}")
            self.already_saved_kato.add(kato_code)

        print("\n🎉 Все файлы сохранены в итоговые_файлы/ОБЛАСТЬ/")
